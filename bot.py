import os
import gc
import sys
import json
import time
import glob
import wave
import queue
import torch
import config
import librosa
import discord
import logging
import asyncio
import hashlib
import tempfile
import datetime
import warnings
import unicodedata
import configparser
import concurrent.futures
from openpyxl import Workbook, load_workbook

warnings.filterwarnings(action='ignore', module='.*paramiko.*')

modules_to_ignore = ["paramiko", "xformers", "fairseq", "discord.client", "discord.gateway", "discord.voice_client", "discord.player", "asyncio", "numba", "ffmpeg"]
for logger_name in modules_to_ignore: logging.getLogger(logger_name).setLevel(logging.ERROR)

if not os.path.exists('logs'): os.makedirs('logs')

console_formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d | %H:%M:%S')

logging.basicConfig(filename= os.path.join('logs', f'{datetime.datetime.today().strftime("%Y-%m-%d")}.txt'), level=logging.DEBUG, format='%(asctime)s; %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

console_handler = logging.StreamHandler()
console_handler.setFormatter(console_formatter) 

logging.getLogger('').addHandler(console_handler)

configdata = configparser.ConfigParser()
configdata.read('config.ini')

try:
    discord_token = configdata.get('discord', 'token')
    bot_activity = configdata.get('discord', 'activity')
    bot_type_activity = configdata.get('discord', 'type_activity')
    default_language = configdata.get('discord', 'language')
    tts_voice = configdata.get('edge_tts', 'voice')
    apikey = configdata.get('elevenlabs', 'api_key')
    modelid = configdata.get('elevenlabs', 'model_id')
    tts_type = configdata.get('tts', 'type_tts')
except configparser.NoSectionError:
    logging.info("Error loading configuration")
    sys.exit(1)

if not os.path.exists('bot_activity.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Event", "Server ID", "Server name", "Status"])
    wb.save('bot_activity.xlsx')

if tts_type == "elevenlabs": import requests
else: import edge_tts
from vc_infer_pipeline import VC
from fairseq import checkpoint_utils
from lib.infer_pack.models import (SynthesizerTrnMs256NSFsid, SynthesizerTrnMs256NSFsid_nono, SynthesizerTrnMs768NSFsid, SynthesizerTrnMs768NSFsid_nono)

try:
    with open('user_languages.json', 'r') as f: user_languages = json.load(f)
except FileNotFoundError: user_languages = {}

def load_language_data(user_id):
    try:
        with open('user_languages.json', 'r') as f: user_languages = json.load(f)
    except FileNotFoundError: user_languages = {}

    try:  user_language = user_languages[str(user_id)]
    except KeyError: user_language = default_language

    with open(os.path.join(os.path.dirname(__file__), "locales", f"{user_language}.json"), "r", encoding="utf-8") as lang_file: return json.load(lang_file)

language_data = load_language_data("default")

os.system('cls' if os.name == 'nt' else 'clear')
logging.info("┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
logging.info("┃  VoiceMe!  -  @impavloh  -  Beta  ┃")  
logging.info("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
logging.info("Loading configuration")

config = config.Config()

intents = discord.Intents.default()
intents.members = True
client = discord.Client(intents=intents)
tree = discord.app_commands.CommandTree(client)

tts_queue = queue.Queue()
is_playing_audio = False

afk_times = {}
user_voices = {}
audio_locks = {}
cooldownData = {}

def file_checksum(file_path):
    with open(file_path, 'rb') as f:
        file_data = f.read()
        return hashlib.md5(file_data).hexdigest()

def get_existing_model_info(category_directory):
    model_info_path = os.path.join(category_directory, 'model_info.json')
    if os.path.exists(model_info_path):
        with open(model_info_path, 'r') as f: return json.load(f)
    return None

def get_model_files(model_path): return [f for f in os.listdir(model_path) if f.endswith('.pth') or f.endswith('.index')]

def should_regenerate_model_info(existing_model_info, model_name, pth_checksum, index_checksum):
    if existing_model_info is None or model_name not in existing_model_info: return True
    return (existing_model_info[model_name]['model_path_checksum'] != pth_checksum or existing_model_info[model_name]['index_path_checksum'] != index_checksum)

def gather_model_info(category_directory, model_name, model_path, existing_model_info):
    model_files = get_model_files(model_path)
    if len(model_files) != 2: return None, False

    pth_file = [f for f in model_files if f.endswith('.pth')][0]
    index_file = [f for f in model_files if f.endswith('.index')][0]
    pth_checksum = file_checksum(os.path.join(model_path, pth_file))
    index_checksum = file_checksum(os.path.join(model_path, index_file))
    regenerate = should_regenerate_model_info(existing_model_info, model_name, pth_checksum, index_checksum)

    return {"title": model_name, "model_path": pth_file, "feature_retrieval_library": index_file, "model_path_checksum": pth_checksum, "index_path_checksum": index_checksum}, regenerate

def generate_model_info_files(user_id=None):
    if user_id is None: language_data = load_language_data(default_language)
    else: language_data = load_language_data(user_id)

    logging.info("Generating model information files")
    folder_info = {}
    model_directory = "models/"
    for category_name in os.listdir(model_directory):
        category_directory = os.path.join(model_directory, category_name)
        if not os.path.isdir(category_directory): continue

        folder_info[category_name] = {"title": category_name, "folder_path": category_name}
        existing_model_info = get_existing_model_info(category_directory)
        model_info = {}
        regenerate_model_info = False

        for model_name in os.listdir(category_directory):
            model_path = os.path.join(category_directory, model_name)
            if not os.path.isdir(model_path): continue

            model_data, regenerate = gather_model_info(category_directory, model_name, model_path, existing_model_info)
            if model_data is not None:
                model_info[model_name] = model_data
                regenerate_model_info |= regenerate

        if regenerate_model_info:
            with open(os.path.join(category_directory, 'model_info.json'), 'w') as f: json.dump(model_info, f, indent=4)

    folder_info_path = os.path.join(model_directory, 'folder_info.json')
    with open(folder_info_path, 'w') as f: json.dump(folder_info, f, indent=4)

generate_model_info_files()

allowed_voices = {}
options = []

model_info_files = glob.glob('models/**/model_info.json', recursive=True)
for model_info_file in model_info_files:
    with open(model_info_file, 'r') as f: model_info = json.load(f)
    for model_name in model_info:
        allowed_voices[model_name] = (model_name, model_name)
        options.append(discord.SelectOption(label=model_name, emoji='📦'))

def create_vc_fn(model_title, tgt_sr, net_g, vc, if_f0, version, file_index):
    def vc_fn(tts_text):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".mp3") as tmpfile:
                tts_mp3_filename = tmpfile.name
                if tts_type == "elevenlabs":
                    url = "https://api.elevenlabs.io/v1/text-to-speech/VR6AewLTigWG4xSOukaG"
                    d = {"text": tts_text, "model_id": modelid, "voice_settings": { "stability": 0.75, "similarity_boost": 1}}
                    h = {"Accept": "audio/mpeg", "Content-Type": "application/json", "xi-api-key": apikey}
                    response = requests.post(url, json=d, headers=h)
                    tmpfile.write(response.content)
                    tmpfile.seek(0)
                else: asyncio.run(edge_tts.Communicate(tts_text, "-".join(tts_voice.split('-')[:-1])).save(tts_mp3_filename))
                audio, sr = librosa.load(tts_mp3_filename, sr=16000, mono=True)
            audio_opt = vc.pipeline(hubert_model, net_g, 0, audio, tts_mp3_filename, [0, 0, 0], -1, "pm", file_index, 0.7, if_f0, 3, tgt_sr, 0, 1, version, 0.5, f0_file=None)
            os.remove(tts_mp3_filename)
            return None if audio_opt is None else (tgt_sr, audio_opt)
        except Exception: return None
    return vc_fn

def load_specific_model(target_category_name, target_model_name):
    global categories
    categories = []
    with open("models/folder_info.json", "r", encoding="utf-8") as f: 
        folder_info = json.load(f)
    for category_name, category_info in folder_info.items():
        if category_name != target_category_name: continue
        category_title = category_info['title']
        category_folder = category_info['folder_path']
        with open(f"models/{category_folder}/model_info.json", "r", encoding="utf-8") as f: 
            models_info = json.load(f)
        for character_name, info in models_info.items():
            if character_name != target_model_name: continue
            model = load_model_data(category_folder, character_name, info)
            categories.append((category_title, category_folder, model))
    del folder_info, models_info
    gc.collect()

    return categories

def load_model_data(category_folder, character_name, info):
    model_title = info['title']
    model_name = info['model_path']
    model_index = f"models/{category_folder}/{character_name}/{info['feature_retrieval_library']}"
    cpt = torch.load(f"models/{category_folder}/{character_name}/{model_name}", map_location="cpu")
    tgt_sr = cpt["config"][-1]
    if_f0 = cpt.get("f0", 1)
    version = cpt.get("version", "v1")
    net_g = load_model_net_g(cpt, cpt.get("f0", 1), cpt.get("version", "v1"))
    del cpt
    net_g.load_state_dict(torch.load(f"models/{category_folder}/{character_name}/{model_name}", map_location="cpu")["weight"], strict=False)
    net_g.eval().to(config.device)
    net_g = net_g.half() if config.is_half else net_g.float()
    vc = VC(tgt_sr, config)
    logging.info(f"RVC ({version}) {character_name} model loaded")
    return [(character_name, model_title, version.upper(), create_vc_fn(model_title, tgt_sr, net_g, vc, if_f0, version, model_index))]

def load_model_net_g(cpt, if_f0, version):
    if version == "v1":
        if if_f0 == 1: net_g = SynthesizerTrnMs256NSFsid(*cpt["config"], is_half=config.is_half)
        else: net_g = SynthesizerTrnMs256NSFsid_nono(*cpt["config"])
    elif version == "v2":
        if if_f0 == 1: net_g = SynthesizerTrnMs768NSFsid(*cpt["config"], is_half=config.is_half)
        else: net_g = SynthesizerTrnMs768NSFsid_nono(*cpt["config"])
    return net_g

def load_hubert():
    global hubert_model
    models, _, _ = checkpoint_utils.load_model_ensemble_and_task(["hubert_base.pt"], suffix="")
    hubert_model = models[0]
    hubert_model = hubert_model.to(config.device)
    hubert_model = hubert_model.half() if config.is_half else hubert_model.float()
    hubert_model.eval()

def remove_special_characters(text): return ''.join(c for c in text if not unicodedata.category(c).startswith('So'))

async def play_audio(voice_client, output_filename):
    global is_playing_audio, tts_queue
    audio_started = asyncio.Future()

    def after_play(_):
        voice_client.stop()
        audio_started.set_result(None)

    while not tts_queue.empty():
        output_filename, audio_data, sample_rate = tts_queue.get()
        audio_source = discord.FFmpegPCMAudio(executable="ffmpeg", source=output_filename)
        audio_source = discord.PCMVolumeTransformer(audio_source, volume=0.7)
        voice_client.play(audio_source, after=after_play)
        num_frames = len(audio_data) // 2
        duration = num_frames / sample_rate
        await asyncio.sleep(duration + 1)

    is_playing_audio = False
    return audio_started

async def run_in_executor(func, *args):
    loop = asyncio.get_event_loop()
    if asyncio.iscoroutinefunction(func): return await func(*args)
    with concurrent.futures.ThreadPoolExecutor() as executor: result = await loop.run_in_executor(executor, func, *args)
    return result

async def get_vc_fn_result(vc_fn, text): return await run_in_executor(vc_fn, text)

load_hubert()
logging.info("Starting bot")

def cooldown(author):
    lastCreation = cooldownData.get(author, time.time() - 5)
    if lastCreation > time.time() - 5:
       timeToWait = 5 - int(time.time() - lastCreation)
       return timeToWait
    cooldownData[author] = time.time()
    return False

class BotonesTTS2(discord.ui.View):
    @discord.ui.button(label="▶️", style=discord.ButtonStyle.green)
    async def reanudar(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        interaction.guild.voice_client.resume()
        await interaction.edit_original_response(view=BotonesTTS(), embed=discord.Embed(title=language_data["tts_playing"], color=0X07ce1b, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_continue"]))
    @discord.ui.button(label="⏹️", style=discord.ButtonStyle.red)
    async def detener(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        interaction.guild.voice_client.stop()
        await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_playing"], color=0Xce0743, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_stop"]))

class DownloadButton(discord.ui.Button):
    def __init__(self, filename, **kwargs):
        super().__init__(style=discord.ButtonStyle.grey, label="⬇️", **kwargs)
        self.filename = filename

    async def callback(self, interaction: discord.Interaction):
        with open(self.filename, 'rb') as fp:
            file = discord.File(fp, filename="VoiceMe-Generated-Audio.wav")
            await interaction.response.send_message(embed=discord.Embed(title=language_data['tts_downloaded_audio'], color=0XBABBE1), file=file)

class BotonesTTS(discord.ui.View):
    def __init__(self, output_filename, **kwargs):
        super().__init__(**kwargs)
        self.output_filename = output_filename

    @discord.ui.button(label="⏸️", style=discord.ButtonStyle.green)
    async def pausar(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        interaction.guild.voice_client.pause()
        await interaction.edit_original_response(view=BotonesTTS2(), embed=discord.Embed(title=language_data["tts_playing"], color=0Xce6307, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_pause"]))
    @discord.ui.button(label="⏹️", style=discord.ButtonStyle.red)
    async def detener(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        interaction.guild.voice_client.stop()
        await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_playing"], color=0Xce0743, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_stop"]))
    @discord.ui.button(label="⬇️", style=discord.ButtonStyle.grey)
    async def descargar(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        if self.output_filename:
            with open(self.output_filename, 'rb') as fp:
                file = discord.File(fp, filename="VoiceMe-Generated-Audio.wav")
                await interaction.followup.send(file=file, embed=discord.Embed(title=language_data["tts_downloaded_audio"], color=0XBABBE1, timestamp=datetime.datetime.now()))
        else: await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_playing"], color=0Xce0743, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_no_audio"]))        

class BotonesTTS3(discord.ui.View):
    def __init__(self, output_filename, **kwargs):
        super().__init__(**kwargs)
        self.output_filename = output_filename

    @discord.ui.button(label="⬇️", style=discord.ButtonStyle.grey)
    async def descargar(self, interaction: discord.Interaction, button: discord.ui.Button):
        language_data = load_language_data(interaction.user.id)
        await interaction.response.defer()
        if self.output_filename:
            with open(self.output_filename, 'rb') as fp:
                file = discord.File(fp, filename="VoiceMe-Generated-Audio.wav")
                await interaction.followup.send(file=file, embed=discord.Embed(title=language_data["tts_downloaded_audio"], color=0XBABBE1, timestamp=datetime.datetime.now()))
        else: await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_playing"], color=0Xce0743, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_no_audio"]))        
                      
class LanguageDropdown(discord.ui.Select):
    def __init__(self):
        options = [discord.SelectOption(label="English", emoji='🇺🇸', value="en"), discord.SelectOption(label="Español", emoji='🇪🇸', value="es"), discord.SelectOption(label="Français", emoji='🇫🇷', value="fr"), discord.SelectOption(label="Deutsch", emoji='🇩🇪', value="de"), discord.SelectOption(label="Português", emoji='🇵🇹', value="pt")]
        super().__init__(placeholder=language_data['language_select'] , min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        global user_languages, user_voices, language_data, modelid
        user_languages[str(interaction.user.id)] = interaction.data['values'][0]
        language_data = load_language_data(interaction.user.id)

        with open('user_languages.json', 'w') as f: json.dump(user_languages, f)

        language_voice_mapping = { "es": "es-ES-AlvaroNeural-Male", "pt": "pt-PT-DuarteNeural-Male", "de": "de-AT-JonasNeural-Male", "fr": "fr-FR-HenriNeural-Male", "en": "en-US-GuyNeural-Male" }

        user_id = str(interaction.user.id)
        user_language = user_languages[user_id]

        if user_language in language_voice_mapping:
            user_voices[user_id] = language_voice_mapping[user_language]
            if user_language == "en": modelid = 'eleven_monolingual_v1'
            else: modelid = 'eleven_multilingual_v1'
        
        await interaction.response.send_message(embed=discord.Embed(title=language_data["language_changed"].format(new_language=user_languages[str(interaction.user.id)]), color=0XBABBE1), ephemeral=True)

class CommandDropdownView(discord.ui.View):
    def __init__(self):
        super().__init__()
        self.add_item(Dropdown())

class Dropdown(discord.ui.Select):
    def __init__(self): super().__init__(placeholder=language_data['choose_voice'], min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        global user_voices
        language_data = load_language_data(interaction.user.id)
        selected_voice = self.values[0]

        if selected_voice in allowed_voices:
            target_category_name, target_model_name = allowed_voices[selected_voice]
            user_voices[interaction.user.id] = (target_category_name, target_model_name)
            logging.info(f"Voice model changed to {selected_voice} ({interaction.user.id} | @{interaction.user})")
            
            embed = discord.Embed(title=language_data['voice_changed'].format(selected_voice=selected_voice) , color=0XBABBE1)
        else: embed = discord.Embed(title=language_data['voice_not_valid'] , color=0XBABBE1)
        await interaction.response.send_message(embed=embed, ephemeral=True)

class CommandDropdownView2(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)
        self.add_item(Dropdown2())

class Dropdown2(discord.ui.Select):
    def __init__(self): 
        super().__init__(placeholder=language_data['choose_voice'], min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        global user_voices, selected_voice
        language_data = load_language_data(interaction.user.id)
        retry_after = cooldown(interaction.user.id)
        if retry_after:
            await interaction.response.send_message(embed=discord.Embed(title=language_data["cooldown"].format(cooldown=int(retry_after)), color=0XBABBE1), ephemeral=True)
            logging.info(f"Cooldown | {interaction.user.id} | @{interaction.user} | Server {interaction.guild.name}")
            return
        
        selected_voice = self.values[0]
        target_category_name, target_model_name = allowed_voices[selected_voice]
        user_voices[interaction.user.id] = (target_category_name, target_model_name)
        load_specific_model(target_category_name, target_model_name)
        logging.info(f"Voice model {selected_voice} loaded ({interaction.user.id} | @{interaction.user})")
        await play_audio_for_interaction(interaction)

async def play_audio_for_model(interaction, voice_client, lock, folder_title, model):
    character_name, model_title, model_version, vc_fn = model
    logging.info(f"Processing TTS with the voice of {folder_title}")

    try:
        async with lock:
            sample_rate, audio_data = await get_vc_fn_result(vc_fn, sanitized_text)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as tmpfile: output_filename = tmpfile.name
            with wave.open(output_filename, "wb") as wav_file:
                wav_file.setnchannels(1)
                wav_file.setsampwidth(2)
                wav_file.setframerate(sample_rate)
                wav_file.writeframes(audio_data)
                wav_file.close()
            tts_queue.put((output_filename, audio_data, sample_rate))
            await play_or_queue_audio(interaction, voice_client, output_filename)
    except Exception as e:
        await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_error"], color=0X990033))
        await voice_client.disconnect()
        logging.error(f"TTS ERROR: {str(e)} | {interaction.user.id} | @{interaction.user} | Server {interaction.guild.name}")

async def play_audio_for_interaction(interaction):
    language_data = load_language_data(interaction.user.id)
    logging.info(f"User {interaction.user} used 'say' command in server {interaction.guild.name}")
    try: await interaction.response.send_message(embed=discord.Embed(title=language_data["tts_generating"], color=0XBABBE1, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_generating2"]), ephemeral=True)
    except discord.errors.NotFound: await interaction.followup.send(embed=discord.Embed(title=language_data["tts_generating"], color=0XBABBE1, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_generating2"]), ephemeral=True)
    voice_client = next((vc for vc in client.voice_clients if vc.guild == interaction.guild), None)
    if voice_client is None: voice_client = await interaction.user.voice.channel.connect(self_deaf=True, self_mute=False)
    lock = audio_locks.setdefault(interaction.guild.id, asyncio.Lock())

    for (folder_title, folder, models) in categories:
        for model in models:
            await play_audio_for_model(interaction, voice_client, lock, folder_title, model)
            
async def play_or_queue_audio(interaction, voice_client, output_filename):
    if not voice_client.is_playing() and not voice_client.is_paused():
        await interaction.edit_original_response(embed=discord.Embed(title=language_data["tts_playing"], color=0XBABBE1), view=BotonesTTS(output_filename=output_filename))
        audio_started = await play_audio(voice_client, output_filename)
        await audio_started
        if not voice_client.is_playing() and not voice_client.is_paused():
            await interaction.edit_original_response(view=BotonesTTS3(output_filename=output_filename), embed=discord.Embed(title=language_data["tts_played"], color=0XBABBE1, timestamp=datetime.datetime.now()))
            logging.info(f"TTS processed and played correctly ({interaction.user.id} | @{interaction.user})")
    else:
        queue_position = tts_queue.qsize() - 1
        await interaction.edit_original_response(view=None, embed=discord.Embed(title=language_data["tts_added_to_queue"], color=0XBABBE1, timestamp=datetime.datetime.now()).set_footer(text=language_data["tts_added_to_queue2"].format(queue_position=queue_position)))
        logging.info(f"TTS added to the queue at position {queue_position} ({interaction.user.id} | @{interaction.user})")

def handle_interaction_exception(interaction, voice_client):
    logging.error(f"Unable to process TTS for user {interaction.user.id} | @{interaction.user}", exc_info=True)
    if voice_client.is_connected():
        if voice_client.is_playing() or voice_client.is_paused():
            voice_client.stop()
        asyncio.ensure_future(voice_client.disconnect(force=True))
        
async def afk_task():
    global afk_times
    while True:
        now = datetime.datetime.utcnow()
        for guild in client.guilds:
            for voice_channel in guild.voice_channels:
                if len(voice_channel.members) == 1:
                    if voice_channel not in afk_times:
                        afk_times[voice_channel] = now
                    elif (now - afk_times[voice_channel]).total_seconds() >= 30:
                        voice_client = voice_channel.guild.voice_client
                        if voice_client:
                            await voice_client.disconnect()
                            del afk_times[voice_channel]
                elif voice_channel in afk_times:
                    del afk_times[voice_channel]
        await asyncio.sleep(1)
            
@client.event
async def on_guild_join(guild):
    wb = load_workbook('bot_activity.xlsx')
    ws = wb.active
    ws.append([datetime.datetime.now(), "Bot joined", guild.id, guild.name, "Online"])
    wb.save('bot_activity.xlsx')
    logging.info("·····································")
    logging.info(f"Bot joined new server {guild.name} | {str(guild.id)}")
    logging.info(f"Now connected to {len(client.guilds)} servers")
    logging.info("·····································")
    await tree.sync()
    
@client.event
async def on_guild_remove(guild):
    wb = load_workbook('bot_activity.xlsx')
    ws = wb.active
    ws.append([datetime.datetime.now(), "Bot left", guild.id, guild.name, "Offline"])
    wb.save('bot_activity.xlsx')
    logging.info("·····································")
    logging.info(f"Bot left a server {guild.name} | {str(guild.id)}")
    if len(client.guilds) == 0: logging.info("Bot isn't in any server.")
    else: logging.info(f"Now connected to {len(client.guilds)} servers")
    logging.info("·····································")
    
@client.event
async def on_ready():
    logging.info("·····································")
    wb = load_workbook('bot_activity.xlsx')
    ws = wb.active
    if len(client.guilds) == 0: logging.info("Bot isn't in any server.")
    else:
        logging.info(f"Bot is connected to {len(client.guilds)} servers:")
        for guild in client.guilds:
            logging.info(f"{guild.name} | {guild.id}")
            if not any(str(guild.id) in str(cell.value) for row in ws.iter_rows() for cell in row): ws.append([datetime.datetime.now(), "Bot already in", str(guild.id), guild.name, "Online"])
    wb.save('bot_activity.xlsx')
    logging.info("·····································")
    await tree.sync()
    language_data = load_language_data(client.user.id)
    logging.info(f"Bot online as {client.user.name}")
    asyncio.create_task(afk_task())
    await client.change_presence(status=discord.Status.online, activity=discord.Game(name=bot_activity, type=bot_type_activity))

@tree.error
async def voz_error(interaction: discord.Interaction, error):
    if isinstance(error, discord.app_commands.errors.CommandOnCooldown):
        await interaction.response.send_message(embed=discord.Embed(title=language_data["cooldown"].format(cooldown=int(error.retry_after)), color=0XBABBE1), ephemeral=True)
        logging.info(f"Cooldown | {interaction.user.id} | @{interaction.user} | Server {interaction.guild.name}")
                
@tree.command(name="join", description="Connects the bot to your voice channel")
async def conectar(interaction: discord.Interaction):
    try:
        language_data = load_language_data(interaction.user.id)
        if interaction.user.voice is not None:
            voice_channel = interaction.user.voice.channel
            if interaction.guild.voice_client is None:
                await voice_channel.connect(self_deaf=True, self_mute=False)
                await interaction.response.send_message(embed=discord.Embed(title=language_data["bot_connected_to_voice_channel"].format(voice_channel=voice_channel), color=0XBABBE1), ephemeral=True)
            elif interaction.guild.voice_client.channel != voice_channel:
                await interaction.guild.voice_client.disconnect()
                await voice_channel.connect(self_deaf=True, self_mute=False)
                await interaction.response.send_message(embed=discord.Embed(title=language_data["bot_moved_to_voice_channel"].format(voice_channel=voice_channel), color=0XBABBE1), ephemeral=True)
            else: await interaction.response.send_message(embed=discord.Embed(title=language_data["bot_already_in_voice_channel"].format(voice_channel=voice_channel), color=0XBABBE1), ephemeral=True)
        else: await interaction.response.send_message(embed=discord.Embed(title=language_data["user_not_in_voice_channel"], color=0xBABBE1), ephemeral=True)
    except Exception as e:
        logging.error(f"Join command error: {str(e)} | {interaction.user.id} | @{interaction.user} | Server {interaction.guild.name}")
        
@tree.command(name="leave", description="Disconnects the bot from the voice channel")
async def salir(interaction: discord.Interaction):
    try:
        language_data = load_language_data(interaction.user.id)
        voice_client = interaction.guild.voice_client
        if not voice_client:
            await interaction.response.send_message(embed=discord.Embed(title=language_data['bot_not_in_voice_channel'], color=0XBABBE1), ephemeral=True)
            return
        await voice_client.disconnect()
        await interaction.response.send_message(embed=discord.Embed(title=language_data['bot_disconnected_from_voice_channel'], color=0XBABBE1), ephemeral=True)
    except Exception as e:
        logging.error(f"Leave command error: {str(e)} | {interaction.user.id} | @{interaction.user} | Server {interaction.guild.name}")

@tree.command(name="help", description="Displays help information")
async def ayuda(interaction: discord.Interaction):
    global user_languages, language_data
    language_data = load_language_data(interaction.user.id)
    embed = discord.Embed(title="🎤 VoiceMe! 🔊 ", color=0XBABBE1, timestamp=datetime.datetime.now())
    embed.set_thumbnail(url="https://www.pavloh.com/static/media/vmlogo.787d69a9ceff9714192d.webp")
    fields = [
            (f":loud_sound: {language_data['dialog_tts_command']}", f"`/say <message>`: {language_data['tts_command_description']}", False),
            (f":speaker: {language_data['dialog_voice_command']}", f"`/join`: {language_data['command_connect_description']}\n" f"`/leave`: {language_data['command_disconnect_description']}", False),
            (f":microphone2: {language_data['dialog_change_voice']}", f"`/voice`: {language_data['voice_command_description']}", False),
            (f":question: {language_data['dialog_extra_commands']}", f"`/help`: {language_data['help_command_description']}", False),
            (f":information_source: {language_data['help_servers'].format(lenservers=len(client.guilds))}", "", True)
        ]
    for name, value, inline in fields:
        embed.add_field(name=name, value=value, inline=inline)
        embed.set_footer(text="VoiceMe! • @impavloh ")
        buttons = [
            discord.ui.Button(label='🐦 Twitter', style=discord.ButtonStyle.link, url='https://x.com/impavloh'),
            discord.ui.Button(label='🌐 Web', style=discord.ButtonStyle.link, url='https://voiceme.pavloh.com'),
            discord.ui.Button(label='☕ Support', style=discord.ButtonStyle.link, url='https://www.buymeacoffee.com/pavloh')
        ]
        view = discord.ui.View()
        for button in buttons:
            view.add_item(button)
    await interaction.response.send_message(embed=embed, view=view)
        
@tree.command(name="language", description="Changes the current language of the bot")
@discord.app_commands.checks.cooldown(1, 5.0, key=lambda i: (i.guild_id, i.user.id))
async def language(interaction: discord.Interaction):
    view = discord.ui.View()
    view.add_item(LanguageDropdown())
    await interaction.response.send_message(view=view, ephemeral=True)
    
@tree.command(name="voice", description="Changes the final voice model of the TTS")
@discord.app_commands.checks.cooldown(1, 5.0, key=lambda i: (i.guild_id, i.user.id))
async def voz(interaction: discord.Interaction):
    global current_voice
    language_data = load_language_data(interaction.user.id)
    current_voice = user_voices.get(interaction.user.id)
    if current_voice is None:  await interaction.response.send_message(embed=discord.Embed(title=language_data["voice_not"], color=0XBABBE1), view=CommandDropdownView(), ephemeral=True)
    else: 
        voice_name = current_voice[1]
        await interaction.response.send_message(embed=discord.Embed(title=language_data["current_voice"].format(current_voice=voice_name), color=0XBABBE1), view=CommandDropdownView(), ephemeral=True)

@tree.command(name="say", description="Speak a message in the voice channel")
@discord.app_commands.checks.cooldown(1, 10.0, key=lambda i: (i.guild_id, i.user.id))
async def tts(interaction: discord.Interaction, mensaje: str):
    if not interaction.user.voice:
        await interaction.response.send_message(embed=discord.Embed(title=language_data["not_in_voice_channel"], color=0XBABBE1), ephemeral=True)
        return
    
    global sanitized_text, user_voices
    sanitized_text = remove_special_characters(mensaje)
    
    user_voice = user_voices.get(interaction.user.id)
    if user_voice is None: await interaction.response.send_message(view=CommandDropdownView2(), embed=discord.Embed(title=language_data["choose_voice"], color=0XBABBE1), ephemeral=True)
    else:
        target_category_name, target_model_name = user_voice
        load_specific_model(target_category_name, target_model_name)
        await play_audio_for_interaction(interaction)
        
client.run(discord_token)
